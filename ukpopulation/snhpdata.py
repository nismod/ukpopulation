"""
snhpdata.py - household projections
"""
import os.path
import zipfile
import pandas as pd
import requests
import tempfile
from openpyxl import load_workbook
import ukpopulation.utils as utils
import pyexcel

class SNHPData:
  """
  Functionality for downloading and collating UK Subnational Household Projection (SNHP) data
  """  

  def __init__(self, cache_dir=utils.default_cache_dir()):
    self.cache_dir = cache_dir

    self.data = {}
    self.data[utils.EN] = self.__do_england()
    self.data[utils.WA] = self.__do_wales()
    self.data[utils.SC] = self.__do_scotland()
    self.data[utils.NI] = self.__do_nireland()

  def unified(self):
    """ 
    Creates a unified dataset containing values for
    - the year range present in all datasets (2016-2039 at time of typing)
    - a lowest-common denimator set of household types (that maps to OA-level census categories)
    """
    raise NotImplementedError("The categories used in each country have no clear common denominator")
    # TODO best I can do is probably differentiate between single-person, multi-person including children, and multi-person no children households

  def min_year(self, code):
    """
    Returns the first year in the projection, assumes a single LAD or country code
    """
    # convert to country if necessary
    if "0" in code:
      code = utils.country(code)[0]
    return min(self.data[code].PROJECTED_YEAR_NAME.unique().astype(int))

  def max_year(self, code):
    """
    Returns the final year in the projection, assumes a single LAD or country code
    """
    # convert to country if necessary
    if "0" in code:
      code = utils.country(code)[0]
    return max(self.data[code].PROJECTED_YEAR_NAME.unique().astype(int))

  def filter(self, categories, geog_codes, years=None):
    # see unified...
    raise NotImplementedError("The categories used in each country have no clear common denominator")

  def aggregate(self, geog_codes, years=None):
    """ Returns aggregate counts of household for specified geographies and years """

    # convert geog_codes and years to arrays if single values supplied (for isin)
    if isinstance(geog_codes, str):
      geog_codes = [geog_codes]

    countries = utils.country(geog_codes)

    # TODO fix incorrect assumption is that all countries have the same year range 
    years = utils.trim_range(years, self.min_year(countries[0]), self.max_year(countries[0]))

    retval = pd.DataFrame()
    # loop over datasets as needed
    for country in countries:
      # apply filters
      retval = retval.append(self.data[country][(self.data[country].GEOGRAPHY_CODE.isin(geog_codes)) & 
                                               (self.data[country].PROJECTED_YEAR_NAME.isin(years))] \
                            ,ignore_index=True, sort=False)
    return retval.groupby(["GEOGRAPHY_CODE", "PROJECTED_YEAR_NAME"]).sum().reset_index()

  def __do_england(self):
    print("Collating SNHP data for England...")
    england_src = "https://www.ons.gov.uk/file?uri=/peoplepopulationandcommunity/populationandmigration/populationprojections/datasets/householdprojectionsforenglanddetaileddataformodellingandanalysis/2016based/detailedtablesstage1and2.zip"
    england_raw = os.path.join(self.cache_dir, os.path.basename(england_src))
    england_processed = self.cache_dir + "/snhp_e.csv"

    if os.path.isfile(england_processed): 
      snhp_e = pd.read_csv(england_processed)
    else: 
      response = requests.get(england_src)
      with open(england_raw, 'wb') as fd:
        for chunk in response.iter_content(chunk_size=1024):
          fd.write(chunk)
        print("Downloaded", england_raw)

      # this doesnt work if you directly supply the file in the zip to load_workbook
      # workaround is to extract the file to a tmp dir and load from there
      z = zipfile.ZipFile(england_raw)
      tmpdir = tempfile.TemporaryDirectory().name
      #print(tmpdir)
      z.extract("detailedtablesstage1and2/s2 Households.xlsx", tmpdir)
      sheet = load_workbook(os.path.join(tmpdir, "detailedtablesstage1and2/s2 Households.xlsx"), read_only=True)["Households"]

      raw = utils.read_cell_range(sheet,"A7", "AS32263")
      snhp_e = pd.DataFrame(raw[1:,:], columns=raw[0,:])

      # remove years before 2011 census and switch years from columns to rows
      snhp_e = snhp_e.drop([str(y) for y in range(2001,2011)], axis=1) \
        .melt(id_vars=["CODE", "AREA", "AGE GROUP", "HOUSEHOLD TYPE"]).drop("AREA", axis=1)
      # ensure count is numeric
      snhp_e.value = snhp_e.value.astype(float)
      # remove age categories and standardise column names
      snhp_e = snhp_e.groupby(["CODE", "HOUSEHOLD TYPE", "variable"]).sum().reset_index() \
        .rename({"CODE": "GEOGRAPHY_CODE", 
                "HOUSEHOLD TYPE": "HOUSEHOLD_TYPE", 
                "variable": "PROJECTED_YEAR_NAME", 
                "value": "OBS_VALUE"}, axis=1)

      snhp_e.to_csv(england_processed, index=False)

    return snhp_e

  def __do_wales(self):
    print("Collating SNHP data for Wales...")

    wales_raw = self.cache_dir + "/snhp_w.csv"
    if not os.path.isfile(wales_raw): 
      fields = ['Area_AltCode1','Year_Code','Data','Area_Hierarchy','Variant_Code','Householdtype_ItemName_ENG']
      # StatsWales is an OData endpoint, so select fields of interest
      url = "http://open.statswales.gov.wales/dataset/hous0115?$select={}".format(",".join(fields))
      # use OData syntax to filter P (persons), AllAges (all ages), Area_Hierarchy 596 (LADs)
      url += "&$filter=Variant_Code eq 1 and Area_Hierarchy eq 'W92000004'" #Householdtype_ItemName_ENG
      data = []
      while True:
        print(url)
        r = requests.get(url)
        r.raise_for_status()
        r_data = r.json()
        data += r_data['value']
        if "odata.nextLink" in r_data:
          url = r_data["odata.nextLink"]
        else:
          break
      snhp_w = pd.DataFrame(data)

      # # Remove unwanted and rename wanted columns
      snhp_w = snhp_w.drop(["Area_Hierarchy", "Variant_Code"], axis=1)
      snhp_w = snhp_w.rename(columns={"Area_AltCode1": "GEOGRAPHY_CODE",
                                      "Data": "OBS_VALUE", 
                                      "Householdtype_ItemName_ENG": "HOUSEHOLD_TYPE",
                                      "Year_Code": "PROJECTED_YEAR_NAME"})
      # reinstate + signs that went missing
      snhp_w.HOUSEHOLD_TYPE.replace(['4 person (2 adults, 1 children)', '5 person (No children)', '5 person (2 adults, 1 children)', '5 person (1 adult, 4 children)'],
                                    ['4 person (2+ adults, 1+ children)', '5+ person (No children)', '5+ person (2+ adults, 1+ children)', '5+ person (1 adult, 4+ children)'], 
                                    inplace=True)
      # Drop occupancy and population (for now, might be useful?)
      snhp_w = snhp_w[~snhp_w.HOUSEHOLD_TYPE.isin(["Households", "Projected Private Household Population", "Average Household Size"])]
      snhp_w.to_csv(wales_raw, index=False)

    # this avoids any issues with the index (which was dropped on save)
    snhp_w = pd.read_csv(wales_raw)
    return snhp_w

  def __do_scotland(self):
    print("Collating SNHP data for Scotland...")

    scotland_processed = os.path.join(self.cache_dir, "snhp_s.csv")
    scotland_src = "https://www.nrscotland.gov.uk/files//statistics/household-projections/16/2016-house-proj-detailed-coun-princ.zip"
    scotland_raw = os.path.join(self.cache_dir, os.path.basename(scotland_src)) 

    if os.path.isfile(scotland_processed): 
      snhp_s = pd.read_csv(scotland_processed)
    else:
      response = requests.get(scotland_src)
      with open(scotland_raw, 'wb') as fd:
        for chunk in response.iter_content(chunk_size=1024):
          fd.write(chunk)
        print("Downloaded", scotland_raw)

        lookup = {"Aberdeen City": "S12000033",
                  "Aberdeenshire": "S12000034",
                          "Angus": "S12000041",
                "Argyll and Bute": "S12000035",
              "Scottish Borders": "S12000026",
              "Clackmannanshire": "S12000005",
            "West Dunbartonshire": "S12000039",
          "Dumfries and Galloway": "S12000006",
                    "Dundee City": "S12000042",
                  "East Ayrshire": "S12000008",
            "East Dunbartonshire": "S12000045",
                  "East Lothian": "S12000010",
              "East Renfrewshire": "S12000011",
              "City of Edinburgh": "S12000036",
                        "Falkirk": "S12000014",
                          "Fife": "S12000015",
                  "Glasgow City": "S12000046",
                      "Highland": "S12000017",
                    "Inverclyde": "S12000018",
                    "Midlothian": "S12000019",
                          "Moray": "S12000020",
                "North Ayrshire": "S12000021",
              "North Lanarkshire": "S12000044",
                "Orkney Islands": "S12000023",
              "Perth and Kinross": "S12000024",
                  "Renfrewshire": "S12000038",
              "Shetland Islands": "S12000027",
                "South Ayrshire": "S12000028",
              "South Lanarkshire": "S12000029",
                      "Stirling": "S12000030",
                  "West Lothian": "S12000040",
            "Na h-Eileanan Siar": "S12000013"}

      z = zipfile.ZipFile(scotland_raw)

      snhp_s = pd.DataFrame()
      for filename in z.namelist():
        council_area = filename[36:-4].strip()
        if council_area == "Scotland":
          continue

        chunk = pd.read_csv(z.open(filename, "r"), encoding="latin1", skiprows=3) \
          .dropna(how="all") \
          .drop(["Unnamed: 28", "Unnamed: 29", "Unnamed: 30"], axis=1) \
          .replace(",", "", regex=True) 
        #print(chunk)
        chunk["Unnamed: 1"] = chunk["Unnamed: 1"].str.strip()
        chunk = chunk[chunk["Unnamed: 1"] != "All ages"].drop("Unnamed: 1", axis=1)
        chunk.fillna(method='ffill', inplace=True)
        chunk["Unnamed: 0"] = chunk["Unnamed: 0"].str.strip()
        # NOTE: commas have been removed from ALL cells, so "1 adult, 1+ children" is now "1 adult 1+ children"
        chunk = chunk[chunk["Unnamed: 0"].isin(["1 adult: male", "1 adult: female", "1 adult 1+ children", "2 adults", "2+ adults 1+ children", "3+ adults", "All households"])]
        chunk = pd.melt(chunk, id_vars="Unnamed: 0")
        chunk.value = chunk.value.astype(int)
        chunk = chunk.groupby(["Unnamed: 0", "variable"]).sum().reset_index().rename({"Unnamed: 0": "HOUSEHOLD_TYPE", "variable": "PROJECTED_YEAR_NAME", "value": "OBS_VALUE"}, axis=1)
        chunk.insert(0, "GEOGRAPHY_CODE", lookup[council_area])
        snhp_s = snhp_s.append(chunk, ignore_index=True)
      snhp_s.to_csv(scotland_processed, index=False)

    return snhp_s

  def __do_nireland(self):
    # 1 worksheet per LAD equivalent
    print("Collating SNHP data for Northern Ireland...")
    ni_src = "https://www.nisra.gov.uk/sites/nisra.gov.uk/files/publications/HHP16_LGD2014.xls"
    ni_processed = os.path.join(self.cache_dir, "snhp_ni.csv")
    if os.path.isfile(ni_processed): 
      snhp_ni = pd.read_csv(ni_processed)
    else:
      ni_raw = os.path.join(self.cache_dir, os.path.basename(ni_src))
      response = requests.get(ni_src)
      with open(ni_raw, 'wb') as fd:
        for chunk in response.iter_content(chunk_size=1024):
          fd.write(chunk)

      districts = ["N090000{:02d}".format(i) for i in range(1,12)]

      # convert to temp xlsx
      tmp_xlsx_file = tempfile.NamedTemporaryFile(suffix=".xlsx").name
      #print(tmp_xlsx_file)
      pyexcel.save_book_as(file_name=ni_raw, dest_file_name=tmp_xlsx_file)
      xlsx_ni = load_workbook(tmp_xlsx_file, read_only=True)

      snhp_ni = pd.DataFrame()

      for d in districts:
        raw = utils.read_cell_range(xlsx_ni[d], "A10", "AA15") # omitting Total row
        data = pd.DataFrame(raw[1:,:], columns=raw[0,:]).melt(id_vars="Household Type*") \
          .rename({"Household Type*": "HOUSEHOLD_TYPE", "variable": "PROJECTED_YEAR_NAME", "value": "OBS_VALUE"}, axis=1)

        data.insert(0, "GEOGRAPHY_CODE", d)
        snhp_ni = snhp_ni.append(data, ignore_index=True)
        
      snhp_ni.to_csv(ni_processed, index=False)

    return snhp_ni

    