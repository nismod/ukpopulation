import os.path
import numpy as np
import pandas as pd
import requests
from openpyxl import load_workbook
import pyexcel
import tempfile

def _read_cell_range(worksheet, topleft, bottomright):
  data_rows = []
  for row in worksheet[topleft:bottomright]:
    data_cols = []
    for cell in row:
      data_cols.append(cell.value)
    data_rows.append(data_cols)
  return np.array(data_rows)


def __do_nireland():
  self_cache_dir = ""
  # Niron 
  # (1 worksheet per LAD equivalent)
  print("Collating SNHP data for Northern Ireland...")
  ni_src = "https://www.nisra.gov.uk/sites/nisra.gov.uk/files/publications/HHP16_LGD2014.xls"
  ni_raw = os.path.join(self_cache_dir, "snhp_ni.csv")
  if os.path.isfile(ni_raw): 
    snhp_ni = pd.read_csv(ni_raw)
  else:
    response = requests.get(ni_src)
    with open(ni_raw, 'wb') as fd:
      for chunk in response.iter_content(chunk_size=1024):
        fd.write(chunk)

    districts = ["N090000{:02d}".format(i) for i in range(1,12)]


    # TODO temp file(s)
    tmp_xlsx_file = tempfile.TemporaryFile()
    # convert to temp xlsx
    pyexcel.save_book_as(file_name=tmp_xlsx_file, 
                         dest_file_name=os.path.join(self_cache_dir, "ni_HHP16_LGD2014.xlsx"))
    xlsx_ni = load_workbook(tmp_xlsx_file, read_only=True)

    snhp_ni = pd.DataFrame()

    for d in districts:
      raw = _read_cell_range(xlsx_ni[d], "A10", "AA16")
      data = pd.DataFrame(raw[1:,:], columns=raw[0,:]).melt(id_vars="Household Type*") \
        .rename({"Household Type*": "HOUSEHOLD_TYPE", "variable": "PROJECTED_YEAR_NAME", "value": "OBS_VALUE"}, axis=1)

      data.insert(0, "GEOGRAPHY_CODE", d)
      print(data)
      snhp_ni = snhp_ni.append(data, ignore_index=True)
      
    snhp_ni.to_csv(ni_raw, index=False)
  return snhp_ni

def do_england():
  sheet = load_workbook("./s2 Households.xlsx", read_only=True)["Households"]

  raw = _read_cell_range(sheet,"A7", "AS32263")
  snhp_e = pd.DataFrame(raw[1:,:], columns=raw[0,:])
  #df = pd.DataFrame(a[:,1:], index=a[:,0], columns=['A', 'B','C','D'])

  # remove years before 2011 census and switch years from columns to rows
  snhp_e = snhp_e.drop([str(y) for y in range(2001,2011)], axis=1) \
    .melt(id_vars=["CODE", "AREA", "AGE GROUP", "HOUSEHOLD TYPE"]).drop("AREA", axis=1)
  #print(snhp_e.columns.values)
  # ensure count is numeric
  snhp_e.value = snhp_e.value.astype(float)
  #print(snhp_e.head())
  # remove age categories and standardise column names
  snhp_e = snhp_e.groupby(["CODE", "HOUSEHOLD TYPE", "variable"]).sum().reset_index() \
    .rename({"CODE": "GEOGRAPHY_CODE", 
            "HOUSEHOLD TYPE": "HOUSEHOLD_TYPE", 
            "variable": "PROJECTED_YEAR_NAME", 
            "value": "OBS_VALUE"}, axis=1)
  #print(snhp_e.head())
  #print(snhp_e.HOUSEHOLD_TYPE.unique())

  snhp_e.to_csv("snhp_e.csv", index=False)

print(pd.read_csv("snhp_e.csv").HOUSEHOLD_TYPE.unique())
print(__do_nireland().HOUSEHOLD_TYPE.unique())
